import openpyxl as pyxl
import os
from dotenv import load_dotenv

import app.utils as u

load_dotenv(".env")
from app.logger import log_info, log_error

# logging.basicConfig(level=logging.INFO, filename=os.getenv('LOG_FILE_NAME'),filemode="w")


async def find_user(crit, val):
    log_info('Starting the process of searching for a user in a file',criteria=crit,value=val)
    try:
        wb = pyxl.load_workbook(os.getenv('XLSX_FILE_NAME'))
        sheet = wb.active
        n_col = ''
        match crit:
            case 'email': n_col = 'F' 
            case 'id' : n_col = 'A'
        i = 2
        while sheet[f'{n_col}{i}'].value != None:
            if sheet[f'{n_col}{i}'].value==val:
                log_info("User is finded at file")
                # logging.info(f"User {crit} = {sheet[f'{n_col}{i}'].value} is finded at file")
                finded = {
                    'telegram_id': sheet[f'A{i}'].value,
                    'first_name': sheet[f'B{i}'].value,
                    'second_name': sheet[f'C{i}'].value,
                    'user_name': sheet[f'D{i}'].value,
                    'user_contact': sheet[f'E{i}'].value,
                    'user_email': sheet[f'F{i}'].value,
                    'user_course': sheet[f'I{i}'].value,
                    'user_group': sheet[f'J{i}'].value
                }           
                wb.close()
                return {"status":"f","data":finded}
            i+=1
        wb.close()
        log_info("User isn't finded at file")
        return {"status":"n","data":None}
    except Exception as e:
        template = "Error when find user at file {0}." # details:\n{1!r}
        message = template.format(type(e).__name__) # , e.args
        log_error('Error when search user at file', details=e.args)
        return {"status":"e","data":None}

async def reg_user(data):
    log_info('Starting the process of searching for a user in a file',user_email=data['emaili'])
    try:
        wb = pyxl.load_workbook(os.getenv('XLSX_FILE_NAME'))
        sheet = wb.active
        sheet.append([data['u_id'],
                    data['firstname'],data['lastname'],
                    data['username'],data['contact'],data['emaili'],
                    data['educational_institution_and_level'],data['year_of_education_and_spec'],None,None])
        wb.save(os.getenv('XLSX_FILE_NAME'))
        wb.close()
        log_info("User writed to file")
        return {"status":"r","data":None}
    except Exception as e:
        template = "Error when reg user at file {0}. details:\n{1!r}"
        message = template.format(type(e).__name__, e.args)
        log_error('Error when write user at file', details=e.args)
        return {"status":"e","data":None}


async def get_courses_url_by_id(user_id):
    log_info('Starting the process of searching user courses in a file')
    finded = {'user_course':None}
    try:
        file = await u.get_file()

        wb = pyxl.load_workbook(os.getenv('XLSX_FILE_NAME'))
        sheet = wb.active

        i = 2
        while sheet[f'A{i}'].value != None:
            if sheet[f'A{i}'].value==user_id:
                finded = {
                    'user_course': str(sheet[f'I{i}'].value)
                }           
                wb.close()
                os.remove(os.getenv('XLSX_FILE_NAME'))

                if finded['user_course']!=None:
                    return finded
                else:
                    finded = {'user_course':None}
                    return finded
            i+=1
        wb.close()
        os.remove(os.getenv('XLSX_FILE_NAME'))
        return finded
    except Exception as e:
        template = "Error when get list of courses {0}. details:\n{1!r}"
        message = template.format(type(e).__name__, e.args)
        log_error('Error when searching user courses at file', details=e.args)
        return finded
    
async def create_promo(course_id,use_count):
    import random
    simv = ['a','b','c','d','e','f','g','h','i','j','k','l','m',
            'n','o','p','q','r','s','t','u','v','w','x','y','z',
            'A','B','C','D','E','F','G','H','I','J','K','L','M',
            'N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
            '0','1','2','3','4','5','6','7','8','9']
    length = 12
    promo = ''.join(random.choice(simv) for i in range(length))
    wb = pyxl.load_workbook(os.getenv('XLSX_FILE_NAME'))
    sheet = wb.active
    i = 2
    while sheet[f'L{i}'].value!=None:
        i+=1
    sheet[f'L{i}'] = promo
    sheet[f'M{i}'] = course_id
    sheet[f'N{i}'] = use_count
    sheet[f'O{i}'] = 0
    wb.save(os.getenv('XLSX_FILE_NAME'))
    wb.close()
    log_info('Promo to course {course_id} is generated: {promo}')
    return {'promo':promo,'course':course_id,'use_count':use_count}

async def check_promo(promo):
    wb = pyxl.load_workbook(os.getenv('XLSX_FILE_NAME'))
    sheet = wb.active
    i = 2
    while sheet[f'L{i}'].value!=None:
        if sheet[f'L{i}'].value==promo:
            if int(sheet[f'O{i}'].value)<int(sheet[f'N{i}'].value):
                sheet[f'O{i}']=int(sheet[f'O{i}'].value)+1
                log_info('Promo {promo} is valid and usable, activated course {sheet[f"M{i}"].value')
                wb.save(os.getenv('XLSX_FILE_NAME'))
                wb.close()
                return {'status':'v','course':sheet[f'M{i}'].value}
            else:
                log_info('Promo {promo} is valid but unusable, count of activation over')
                wb.save(os.getenv('XLSX_FILE_NAME'))
                wb.close()
                return {'status':'n','course':None}
        i+=1
    log_info('Promo {promo} is invalid, it does not exists')
    wb.save(os.getenv('XLSX_FILE_NAME'))
    wb.close()
    return {'status':'n','course':None}

async def add_course_to_user(course,user):
    wb = pyxl.load_workbook(os.getenv('XLSX_FILE_NAME'))
    sheet = wb.active
    i = 2
    while sheet[f'A{i}'].value != None:
        if sheet[f'A{i}'].value==user:
            if (sheet[f'I{i}'].value)!= None:
                sheet[f'I{i}']=str(sheet[f'I{i}'].value)+f',{course}'
            else:
                sheet[f'I{i}']=f'{course}'
            wb.save(os.getenv('XLSX_FILE_NAME'))
            wb.close()
            return 
        i+=1
    wb.save(os.getenv('XLSX_FILE_NAME'))
    wb.close()
    return 
    
