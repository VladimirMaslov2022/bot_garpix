import openpyxl as pyxl
import os,logging
from dotenv import load_dotenv

import app.utils as u
import app.req_moodle as req

load_dotenv(".env")

logging.basicConfig(level=logging.INFO, filename=os.getenv('LOG_FILE_NAME'),filemode="w")




async def register_user(data):
    try:
        m = await req.find_user('email',data['emaili'])
        file = await u.get_file()
        if len(m['users'])==0:
            rm = await req.reg_user(data)
            if rm:
                e = await find_user('email',data['emaili'])
                if not e and file!=False:
                    wb = pyxl.load_workbook(os.getenv('XLSX_FILE_NAME'))
                    s = 'ssss'.encode()
                    sheet = wb.active
                    sheet.append([data['u_id'],
                                data['firstname'],data['lastname'],
                                data['username'],data['contact'],data['emaili'],
                                data['educational_institution_and_level'],data['year_of_education_and_spec'],None,None])
                    wb.save(os.getenv('XLSX_FILE_NAME'))
                    wb.close()
                    logging.info("Writed to xlsx")
                    # m = await req.find_user('email',data['emaili'])
                    rf = await u.return_file()
                    if rf: return 'registred'
                    else: return 'error'
                else:
                    logging.warning("User is already exists at file")
                    os.remove(os.getenv('XLSX_FILE_NAME'))
                    return 'exists'
            else:
                logging.warning("Error when reg user at moodle")
                return 'error'
        else:
            logging.warning("User is already exists at moodle")
            e = await find_user('email',data['emaili'])
            file = await u.get_file()
            if not e and file!=False:
                wb = pyxl.load_workbook(os.getenv('XLSX_FILE_NAME'))        
                sheet = wb.active
                sheet.append([data['u_id'],
                            m['users'][0]['firstname'],m['users'][0]['lastname'],
                            m['users'][0]['username'],data['contact'],m['users'][0]['email'],
                            data['educational_institution_and_level'],data['year_of_education_and_spec'],None,None])
                wb.save(os.getenv('XLSX_FILE_NAME'))
                wb.close()
                logging.info("Writed to xlsx")
                rf = await u.return_file()
                if rf: return 'registred'
                else: return 'error'
            else:
                logging.warning("User is already exists at moodle and at file")
                return 'exists'
    except Exception as e:
        if e.args != FileNotFoundError:
            template = "Error when reg user at moodle or write user to file {0}. details:\n{1!r}"
            message = template.format(type(e).__name__, e.args)
            logging.error(message)
            return 'error'

    
async def check_user_groups_and_courses_by_id(user_id):
    try:
        file = await u.get_file()

        wb = pyxl.load_workbook(os.getenv('XLSX_FILE_NAME'))
        sheet = wb.active

        i = 2
        while sheet[f'A{i}'].value != None:
            if sheet[f'A{i}'].value==user_id:
                finded = {
                    'user_email': sheet[f'F{i}'].value,
                    'user_course': sheet[f'I{i}'].value,
                    'user_group': sheet[f'J{i}'].value
                }           
                wb.close()
                os.remove(os.getenv('XLSX_FILE_NAME'))

                if finded['user_course']!=None:
                    r = await req.find_user('email', finded['user_email'])

                    data = {
                        'user_id': r['users'][0]['id'],
                        'user_course': str(finded['user_course']),
                        'user_group': finded['user_group']
                    }            

                    await req.add_user_to_course(data)
                    await req.add_user_to_group(data)
                    return True
                else:
                    return False
            i+=1
        wb.close()
        os.remove(os.getenv('XLSX_FILE_NAME'))
        return False
    except Exception as e:
        template = "Error when add user to course or read user courses {0}. details:\n{1!r}"
        message = template.format(type(e).__name__, e.args)
        logging.error(message)
        return False


async def get_courses_url_by_id(user_id):
    finded = {'user_course':None}
    try:
        file = await u.get_file()

        wb = pyxl.load_workbook(os.getenv('XLSX_FILE_NAME'))
        sheet = wb.active

        i = 2
        while sheet[f'A{i}'].value != None:
            if sheet[f'A{i}'].value==user_id:
                finded = {
                    'user_course': str(sheet[f'I{i}'].value)
                }           
                wb.close()
                os.remove(os.getenv('XLSX_FILE_NAME'))

                if finded['user_course']!=None:
                    return finded
                else:
                    finded = {'user_course':None}
                    return finded
            i+=1
        wb.close()
        os.remove(os.getenv('XLSX_FILE_NAME'))
        return finded
    except Exception as e:
        template = "Error when get list of courses {0}. details:\n{1!r}"
        message = template.format(type(e).__name__, e.args)
        logging.error(message)
        return finded